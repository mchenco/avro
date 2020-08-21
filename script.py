import os
import pandas
import shutil
from datetime import datetime
from openpyxl import load_workbook


def main(dirname):
	df = pandas.read_excel(io="raw_data_1.xlsx", sheet_name='Data')

	# parsing input data
	d1_label = df.iloc[1, 2].split(' ')[0]
	d1 = list(df.iloc[2:, 5])
	d1s1, d1s2, d1s3, d1b1, d1b2, d1b3 = save_data(d1)
	d1b1_avg_list = calc_avg(d1b1)
	d1b2_avg_list = calc_avg(d1b2)
	d1b3_avg_list = calc_avg(d1b3)

	d2_label = df.iloc[0, 2].split(' ')[0]
	d2 = list(df.iloc[2:, 8])
	d2s1, d2s2, d2s3, d2b1, d2b2, d2b3 = save_data(d2)
	d2b1_avg_list = calc_avg(d2b1)
	d2b2_avg_list = calc_avg(d2b2)
	d2b3_avg_list = calc_avg(d2b3)

	# from 48h sheet
	df2 = pandas.read_excel(io="raw_data_2.xlsx", sheet_name='Data')
	d1_48 = list(df2.iloc[2:, 5])
	save_48_data(
		d1_48, d1s1, d1s2, d1s3, d1b1, d1b2, d1b3,
		d1b1_avg_list, d1b2_avg_list, d1b3_avg_list
	)

	d2_48 = list(df2.iloc[2:, 8])
	save_48_data(
		d2_48, d2s1, d2s2, d2s3, d2b1, d2b2, d2b3,
		d2b1_avg_list, d2b2_avg_list, d2b3_avg_list
	)

	# write all data to 3 separate docs
	output_data(
		'%s/output1.xlsx' % dirname, d1_label, d2_label, d1s1, d2s1, d1b1, d2b1,
		d1b1_avg_list, d2b1_avg_list
	)
	output_data(
		'%s/output2.xlsx' % dirname, d1_label, d2_label, d1s2, d2s2, d1b2, d2b2,
		d1b2_avg_list, d2b2_avg_list
	)
	output_data(
		'%s/output3.xlsx' % dirname, d1_label, d2_label, d1s3, d2s3, d1b3, d2b3,
		d1b3_avg_list, d2b3_avg_list
	)


def save_data(data):
	ss = 15  # samplesize
	return (
		data[0:ss], data[ss:ss * 2],
		data[ss * 2:ss * 3], data[ss * 3:ss * 4],
		data[ss * 4:ss * 5], data[ss * 5:ss * 6]
	)


def save_48_data(
	data, s1, s2, s3, b1, b2, b3,
	b1_avg_list, b2_avg_list, b3_avg_list):

	ss = 5
	s1 += data[0:ss]
	s2 += data[ss:ss * 2]
	s3 += data[ss * 2: ss * 3]
	b1 += data[ss * 3: ss * 5]
	b2 += data[ss * 4: ss * 5]
	b3 += data[ss * 5: ss * 6]
	b1_avg_list += [sum(data[ss * 3: ss * 5]) / 5]
	b2_avg_list += [sum(data[ss * 4: ss * 5]) / 5]
	b3_avg_list += [sum(data[ss * 5: ss * 6]) / 5]


# returns list of averages
def calc_avg(data):
	lst = [0, 0, 0]
	for i in range(0, 15, 3):
		lst[0] += data[i]
		lst[1] += data[i + 1]
		lst[2] += data[i + 2]
	return [x / 5 for x in lst]


def output_data(
	output_file, d1_label, d2_label, drug1, drug2, drug1_blank, drug2_blank,
	drug1_blankavg, drug2_blankavg):
	wb = load_workbook(output_file)
	paste_blank_data(wb, d1_label, 'Blank S1', drug1_blank)
	paste_blank_data(wb, d2_label, 'Blank S2', drug2_blank)
	paste_data(wb, d1_label, 'S1', drug1, drug1_blankavg)
	paste_data(wb, d2_label, 'S2', drug2, drug2_blankavg)
	wb.save(output_file)


def paste_blank_data(wb, label, worksheet, data):
	try:
		ws = wb[worksheet]
	except:
		ws = wb[label + ' Blank']
	i = 0
	for row in ws.iter_rows(min_row=5, min_col=8, max_row=9, max_col=10):
		for cell in row:
			cell.value = data[i]
			i += 1
	for col in ws.iter_cols(min_row=5, min_col=11, max_row=9, max_col=11):
		for cell in col:
			cell.value = data[i]
			i += 1
	ws.title = label + ' Blank'


def paste_data(wb, label, worksheet, data, avg_list):
	try:
		ws = wb[worksheet]
	except:
		ws = wb[label + ' Sample']
	i, j = 0, 0
	for row in ws.iter_rows(min_row=5, min_col=8, max_row=9, max_col=10):
		for cell in row:
			if data[i] - avg_list[j] < 0:
				cell.value = data[i]
			else:
				cell.value = data[i] - avg_list[j]
			i += 1
			j = j + 1 if j < 2 else 0
	for col in ws.iter_cols(min_row=5, min_col=11, max_row=9, max_col=11):
		for cell in col:
			if data[i] - avg_list[3] < 0:
				cell.value = data[i]
			else:
				cell.value = data[i] - avg_list[3]
			i += 1
	ws.title = label + ' Sample'


if __name__ == "__main__":
	# raw_data1 = input('Enter the name of the 2-24h raw data file: ')
	# raw_data2 = input('Enter the name of the 48h raw data file: ')
	try:
		dirname = datetime.today().strftime('%Y%m%d') + ' Analysis'
		os.mkdir(dirname)
		shutil.copy2('empty_template.xlsx', '%s/output1.xlsx' % dirname)
		shutil.copy2('empty_template.xlsx', '%s/output2.xlsx' % dirname)
		shutil.copy2('empty_template.xlsx', '%s/output3.xlsx' % dirname)
	except:
		print('Files exist. Appending files instead')
		pass

	main(dirname)
